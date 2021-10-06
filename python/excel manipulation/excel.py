import os, pandas, xlrd, string, math
from openpyxl import Workbook

PATH = "./input/"
PATH_OUT = "./output/"
ALPHABET = string.ascii_lowercase.upper()
ALPHABET_LENGTH = len(ALPHABET)

files = os.listdir(PATH)

def get_column_name_at_index(index):
    div = math.floor(index / ALPHABET_LENGTH)
    mod = index % ALPHABET_LENGTH
    res = ""
    if div > 0:
        res += ALPHABET[div - 1]
    res += ALPHABET[mod]
    return res

for file in files:
    # check if the file is to be parsed
    file_name = os.path.join(PATH, file)
    if ((".xlsx" not in file) and (".xls" in file) and os.path.isfile(file_name)):
        print(f"\n[FILE] found {file}")
        try:
            # open the file
            xls = xlrd.open_workbook(r'' + file_name, on_demand=True)
            # start a copy of the file
            wb = Workbook()
            # get the sheets
            sheets = xls.sheet_names()
            for sheet in sheets:
                print(f"[FILE] found sheet '{sheet}'")
                if "Logfile" in sheet:
                    print(f"[SHEET] reading")
                    # read the sheet
                    data = xls.get_sheet(sheets.index(sheet))
                    number_rows = data.nrows
                    number_columns = data.ncols
                    print(f"[SHEET] found {number_rows} rows and {number_columns} columns")
                    # duplicate the sheet in the new file
                    ws = wb.create_sheet(sheet)
                    # find where the speed and acceleration columns are
                    header = data.row(0)
                    time_index = -1
                    speed_index = -1
                    acceleration_index = -1
                    for i in range(number_columns):
                        if header[i].ctype == 1:
                            val = header[i].value
                            if val.lower() == "speed": speed_index = i
                            if val.lower() == "acceleration": acceleration_index = i
                            if val.lower() == "time": time_index = i
                    time_column = get_column_name_at_index(time_index)
                    speed_column = get_column_name_at_index(speed_index)
                    print(f"[SHEET] Time located at column {time_index} ({time_column}) Speed located at column {speed_index} ({speed_column}), Acceleration at column {acceleration_index}")
                    # create the headers in the new sheet
                    print(f"[SHEET] creating headers")
                    header_new = []
                    for i in range(number_columns):
                        header_new.append(header[i].value)
                    header_new.insert(speed_index + 1, "mean speed")
                    header_new.insert(speed_index + 2, "(mean speed-speed)^2")
                    header_new.insert(speed_index + 3, "[Σ((Mean speed-speed)^2]/(count-1)")
                    header_new.insert(speed_index + 4, "{[Σ((Mean speed-speed)^2]/(count-1)}^(1/2)")
                    for i in range(len(header_new)):
                        cell = ws.cell(row=1, column=i+1)
                        cell.value = header_new[i]
                    # create the data row-by-row, and insert the formulas in the appropriate columns
                    print(f"[SHEET] populating data")
                    for i in range(number_rows - 1):
                        # read the data row-by-row
                        row = data.row(i + 1)
                        row_new = []
                        # map the new data
                        for j in range(number_columns):
                            c_type = row[j].ctype
                            c_value = row[j].value
                            if c_type == 1 or c_type == 2:
                                row_new.append(c_value)
                            else:
                                row_new.append("")
                        index_1 = str(i + 2)
                        index_2 = str(i + 1)
                        row_new.insert(speed_index + 1, "=AVERAGE(" + speed_column + ":" + speed_column + ")")
                        # "(mean speed-speed)^2      =(R2-Q2)^2"
                        new_col_1 = get_column_name_at_index(speed_index + 1)
                        row_new.insert(speed_index + 2, "=(" + new_col_1 + index_1 + "-" + speed_column + index_1 + ")^2")
                        # [Σ((Mean speed-speed)^2]/(count-1)     =SUM(S:S)/(COUNT(S:S)-1)
                        new_col_2 = get_column_name_at_index(speed_index + 2)
                        row_new.insert(speed_index + 3, "=SUM(" + new_col_2 + ":" + new_col_2 + ")/(COUNT(" + new_col_2 + ":" + new_col_2 + ")-1)")
                        # {[Σ((Mean speed-speed)^2]/(count-1)}^(1/2)     =SQRT(T2)
                        new_col_3 = get_column_name_at_index(speed_index + 3)
                        row_new.insert(speed_index + 4, "=SQRT(" + new_col_3 + index_1 + ")")
                        # acceleration      =(Q3-Q2)/(G3-G2)*10^3
                        acceleration_value = ""
                        if i > 0:
                            acceleration_value = "=(" + speed_column + index_1 + "-" + speed_column + index_2 + ")/(" + time_column + index_1 + "-" + time_column + index_2 + ")*10^3"
                        row_new[acceleration_index + 4] = acceleration_value
                        for j in range(len(row_new)):
                            cell = ws.cell(row=i+2, column=j+1, value=row_new[j])
                else:
                    print(f"[SHEET] skipping")
            file_name_new = PATH_OUT + file[0 : file.rfind(".")] + ".xlsx"
            wb.save(file_name_new)
            print(f"[FILE] saved new file: {file_name_new}")
        except Exception as e:
            print(f"[FILE] {file} failed: {e}")
            continue

input("\nPress Enter to exit...")
# run 'pyinstaller excel.py' to create an executable